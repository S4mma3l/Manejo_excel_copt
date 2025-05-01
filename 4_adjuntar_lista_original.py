import pandas as pd
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os

# Función para seleccionar un archivo usando un diálogo de selección de archivos
def seleccionar_archivo(titulo):
    Tk().withdraw()  # Ocultar la ventana principal de tkinter
    archivo = askopenfilename(title=titulo, filetypes=[("Archivos Excel", "*.xlsx")])
    if not archivo:
        print("No se seleccionó ningún archivo.")
        exit()
    return archivo

# Directorio que contiene los archivos Excel fuente
source_dir = r'Z:\Autocad Files\PROYECTOS 2024\RESIDENCIA LOTE 10\Mueble_de_Cocina\Solidos\Materiales'

# Obtener la lista de archivos de Excel en el directorio
archivos_excel = [f for f in os.listdir(source_dir) if f.endswith('.xlsx')]

# Mostrar la lista de archivos y pedir al usuario que elija uno
print("Archivos disponibles para usar como fuente:")
for i, archivo in enumerate(archivos_excel):
    print(f"{i + 1}. {archivo}")

# Solicitar al usuario que seleccione un archivo fuente
try:
    opcion_fuente = int(input("Seleccione el número del archivo fuente que desea usar: ")) - 1
    if opcion_fuente < 0 or opcion_fuente >= len(archivos_excel):
        raise ValueError("Selección no válida.")
except ValueError as e:
    print(e)
    exit()

# Cargar el archivo Excel fuente seleccionado
source_file = os.path.join(source_dir, archivos_excel[opcion_fuente])
try:
    source_data = pd.read_excel(source_file, engine='openpyxl')
except Exception as e:
    print(f"Error al leer el archivo fuente: {e}")
    exit()

# Seleccionar el archivo destino utilizando un diálogo de selección de archivos
destination_file = seleccionar_archivo("Seleccione el archivo destino para cargar los datos")

try:
    wb = load_workbook(destination_file)
    if 'LISTA' not in wb.sheetnames:
        print("La hoja 'lista' no se encuentra en el archivo destino.")
        exit()
    ws = wb['LISTA']
except Exception as e:
    print(f"Error al cargar el archivo destino: {e}")
    exit()

# Insertar los datos a partir de la fila 12
start_row = 13  # Comenzar en la fila 12

for index, row in source_data.iterrows():
    # Insertar datos en la fila correspondiente
    row_num = start_row + index
    ws[f'A{row_num}'] = row.get('QTY', '')            # Cant
    ws[f'B{row_num}'] = row.get('DESCRIPTION', '')    # PIEZA
    ws[f'C{row_num}'] = row.get('Largo', '')          # LARGO
    ws[f'D{row_num}'] = row.get('Ancho', '')          # ANCHO
    ws[f'E{row_num}'] = row.get('Espesor', '')        # GRUESO
    ws[f'F{row_num}'] = ''                           # COD (vacío)
    ws[f'G{row_num}'] = row.get('Material', '')       # MATERIAL
    ws[f'H{row_num}'] = row.get('Archivo', '')         # MODULO

# Guardar el archivo Excel modificado
try:
    wb.save(destination_file)
    print("Datos insertados correctamente en la hoja 'lista'.")
except Exception as e:
    print(f"Error al guardar el archivo destino: {e}")
