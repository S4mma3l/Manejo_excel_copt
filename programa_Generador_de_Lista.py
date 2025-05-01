import pandas as pd
import os
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Directorio que contiene los archivos Excel fuente
directorio = r'Z:\Autocad Files\PROYECTOS 2024\RESIDENCIA LOTE 10\Mueble_de_Cocina\Solidos\Materiales'

# Función para seleccionar un archivo usando un diálogo de selección de archivos
def seleccionar_archivo(titulo):
    Tk().withdraw()  # Ocultar la ventana principal de tkinter
    archivo = askopenfilename(title=titulo, filetypes=[("Archivos Excel", "*.xlsx")])
    if not archivo:
        print("No se seleccionó ningún archivo.")
        exit()
    return archivo

# Función para eliminar "mm" de las celdas de texto en todos los archivos Excel
def eliminar_mm():
    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            df = df.applymap(lambda x: x.replace('mm', '') if isinstance(x, str) else x)
            df.to_excel(ruta_archivo, index=False, engine='openpyxl')
            print(f"Se han eliminado las letras 'mm' de {archivo}.")
        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

# Función para consolidar datos de todos los archivos Excel en un solo archivo
def consolidar_datos():
    columnas_interes = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor']
    df_consolidado = pd.DataFrame()
    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]

    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            columnas_disponibles = [col for col in columnas_interes if col in df.columns]
            df_seleccionado = df[columnas_disponibles].dropna(how='all')
            df_seleccionado['Archivo'] = os.path.splitext(archivo)[0]
            df_consolidado = pd.concat([df_consolidado, df_seleccionado], ignore_index=True)
        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

    columnas_finales = ['Archivo'] + columnas_interes
    columnas_finales = [col for col in columnas_finales if col in df_consolidado.columns]
    df_consolidado = df_consolidado[columnas_finales]
    ruta_salida = os.path.join(directorio, 'Union_Archivos.xlsx')
    try:
        df_consolidado.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"Se ha creado el archivo consolidado: {ruta_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo consolidado: {e}")

# Función para asignar materiales basados en el espesor
def asignar_materiales():
    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]
    print("Archivos disponibles:")
    for i, archivo in enumerate(archivos_excel):
        print(f"{i + 1}. {archivo}")

    try:
        opcion = int(input("Elige el número del archivo que deseas procesar: "))
        if 1 <= opcion <= len(archivos_excel):
            archivo_seleccionado = archivos_excel[opcion - 1]
            ruta_archivo = os.path.join(directorio, archivo_seleccionado)
            mapa_espesores = {16: 'MELAMINA 16 MM', 18: 'EUCALIPTO 18 MM', 20: 'MADERA MELINA', 17: 'MELAMINA 18 MM', 22: 'MADERA MELINA'}
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            if 'Espesor' in df.columns:
                df['Material'] = df['Espesor'].map(mapa_espesores)
            else:
                print("La columna 'Espesor' no se encuentra en el archivo.")
                exit()
            nombre_archivo_modificado = f'Asignacion_Materiales_{archivo_seleccionado}'
            ruta_salida = os.path.join(directorio, nombre_archivo_modificado)
            try:
                df.to_excel(ruta_salida, index=False, engine='openpyxl')
                print(f"Se ha añadido la columna 'Material' y se ha guardado el archivo modificado: {ruta_salida}")
            except Exception as e:
                print(f"Error al guardar el archivo modificado: {e}")
        else:
            print("Opción no válida. Por favor, selecciona un número de la lista.")
    except ValueError:
        print("Entrada no válida. Por favor, ingresa un número entero.")

# Función para insertar datos en una hoja específica
def insertar_datos_en_hoja():
    source_file = seleccionar_archivo("Seleccione el archivo fuente")
    try:
        source_data = pd.read_excel(source_file, engine='openpyxl')
    except Exception as e:
        print(f"Error al leer el archivo fuente: {e}")
        exit()

    destination_file = seleccionar_archivo("Seleccione el archivo destino para cargar los datos")
    try:
        wb = load_workbook(destination_file)
        if 'LISTA' not in wb.sheetnames:
            print("La hoja 'lista' no se encuentra en el archivo destino.")
            exit()
        ws = wb['LISTA']
    except Exception as e:
        print(f"Error al cargar el archivo destino: {e}")
        exit()

    start_row = 13
    for index, row in source_data.iterrows():
        row_num = start_row + index
        ws[f'A{row_num}'] = row.get('QTY', '')
        ws[f'B{row_num}'] = row.get('DESCRIPTION', '')
        ws[f'C{row_num}'] = row.get('Largo', '')
        ws[f'D{row_num}'] = row.get('Ancho', '')
        ws[f'E{row_num}'] = row.get('Espesor', '')
        ws[f'F{row_num}'] = ''
        ws[f'G{row_num}'] = row.get('Material', '')
        ws[f'H{row_num}'] = row.get('Archivo', '')

    try:
        wb.save(destination_file)
        print("Datos insertados correctamente en la hoja 'lista'.")
    except Exception as e:
        print(f"Error al guardar el archivo destino: {e}")

# Llamar a las funciones en el orden deseado
eliminar_mm()
consolidar_datos()
asignar_materiales()
insertar_datos_en_hoja()
