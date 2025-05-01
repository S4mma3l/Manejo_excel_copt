# --- START OF FILE programa_v2.0.py ---

import pandas as pd
import os
from openpyxl import load_workbook, styles
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import webbrowser
import traceback # Para logs de error más detallados

# --- Funciones Auxiliares ---

def seleccionar_carpeta(titulo):
    """Función para seleccionar una carpeta."""
    carpeta = filedialog.askdirectory(title=titulo)
    if not carpeta:
        # No mostrar error aquí, la función que llama decidirá
        # messagebox.showerror("Error", "No se seleccionó ninguna carpeta.")
        return None
    return carpeta

def seleccionar_archivo(titulo, filetypes=[("Archivos Excel", "*.xlsx")]):
    """Función para seleccionar un archivo."""
    archivo = filedialog.askopenfilename(title=titulo, filetypes=filetypes)
    if not archivo:
        # No mostrar error aquí
        # messagebox.showerror("Error", "No se seleccionó ningún archivo.")
        return None
    return archivo

# --- Funciones Internas del Flujo Principal (no requieren botones) ---

def _eliminar_mm_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    """
    Elimina 'mm' de las celdas de texto en archivos Excel de un directorio.
    Retorna True si se completa (incluso si no hay archivos), False si hay errores graves.
    """
    print(f"--- Iniciando eliminación de 'mm' en: {directorio} ---")
    try:
        # Excluir archivos temporales y el propio archivo de fusión
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}")
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}")
        return False

    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para procesar.")
        return True # No es un error si no hay archivos

    errores = []
    modificado_alguno = False
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para eliminar 'mm'...")
            # Usar ExcelFile para leer potencialmente múltiples hojas si fuera necesario
            xls = pd.ExcelFile(ruta_archivo, engine='openpyxl')
            # Usar ExcelWriter para escribir potencialmente múltiples hojas
            # Usamos modo 'a' (append) con if_sheet_exists='replace' para sobrescribir hojas existentes
            # Nota: Esto requiere instalar 'openpyxl' >= 3.0.0
            # Si da problemas, volver a leer todo y escribir con pd.ExcelWriter(ruta...)
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                archivo_modificado_en_hoja = False
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    # Convertir todo a string temporalmente para reemplazo seguro
                    df_str = df.astype(str)
                    if df_str.apply(lambda col: col.str.contains('mm', case=False, na=False)).any().any():
                        # Aplicar reemplazo solo si 'mm' está presente (case-insensitive)
                        # Usamos .map para aplicar la función a cada celda
                        df_mod = df.map(lambda x: x.replace('mm', '').replace('MM', '').strip() if isinstance(x, str) else x)

                        # Solo escribir si hubo cambios reales (comparar antes y después)
                        if not df.equals(df_mod):
                             df_mod.to_excel(writer, sheet_name=sheet_name, index=False)
                             archivo_modificado_en_hoja = True
                             print(f"  'mm' encontrado y eliminado en hoja '{sheet_name}' de {archivo}.")
                        else:
                             # Si no hubo cambios en esta hoja, igual hay que escribirla para no perderla
                             # pero solo si el modo es 'w' (sobrescribir todo el archivo).
                             # Con modo 'a' if_sheet_exists='replace', solo necesitamos escribir las modificadas.
                             # Si usamos el enfoque de leer todo y escribir con pd.ExcelWriter normal, sí necesitaríamos escribirla aquí.
                             # Para simplificar con modo 'a', si no hubo cambio no hacemos nada aquí.
                             # Si no se encontró 'mm' en ninguna celda, df_mod será igual a df.
                             pass
                    # else: # Si no hay 'mm' en la hoja, no necesitamos reescribirla con modo 'a' + replace

                if archivo_modificado_en_hoja:
                    modificado_alguno = True
                else:
                    print(f"  No se encontró 'mm' (o no hubo cambios tras quitarlo) en {archivo}.")
            # El `with` se encarga de cerrar y guardar el writer

        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante eliminación de 'mm': {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(error_msg)
            # Podríamos decidir continuar o parar aquí
            # return False # Descomentar si queremos parar ante el primer error

    if errores:
        messagebox.showwarning("Advertencia Eliminación 'mm'", "Ocurrieron errores al eliminar 'mm' en algunos archivos:\n\n" + "\n".join(errores))
        # Decidimos continuar incluso con errores parciales, pero se podría cambiar
    elif modificado_alguno:
         print("Proceso de eliminación de 'mm' completado con modificaciones.")
    else:
         print("Proceso de eliminación de 'mm' completado. No se realizaron modificaciones.")

    print("--- Finalizada eliminación de 'mm' ---")
    return True # Indicar que el proceso general terminó (aunque pudo haber errores parciales)

def _consolidar_datos_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    """
    Consolida datos de archivos Excel en un directorio especificado.
    Retorna la ruta del archivo consolidado si tiene éxito y datos, None si falla o no hay datos.
    """
    print(f"--- Iniciando consolidación de datos en: {directorio} ---")
    columnas_interes = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor']
    df_consolidado = pd.DataFrame()
    ruta_salida = os.path.join(directorio, archivo_consolidado_nombre)

    try:
        # Excluir archivos temporales y el propio archivo de fusión si ya existe
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}")
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}")
        return None

    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para consolidar.")
        messagebox.showinfo("Consolidación", "No se encontraron archivos Excel (aparte de {} si existe) para consolidar.".format(archivo_consolidado_nombre))
        return None # No se puede consolidar si no hay archivos

    errores = []
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para consolidación...")
            # Leer la primera hoja por defecto
            df = pd.read_excel(ruta_archivo, engine='openpyxl')

            # Verificar columnas presentes
            columnas_disponibles = [col for col in columnas_interes if col in df.columns]
            if not columnas_disponibles:
                print(f"  Advertencia: Ninguna de las columnas de interés {columnas_interes} encontrada en {archivo}. Saltando archivo.")
                continue

            df_seleccionado = df[columnas_disponibles].copy() # Usar .copy()

            # Añadir columnas faltantes del grupo de interés con NA
            for col in columnas_interes:
                if col not in df_seleccionado.columns:
                    df_seleccionado[col] = pd.NA

            # Eliminar filas donde TODAS las columnas de interés son NaN/NA
            df_seleccionado = df_seleccionado.dropna(subset=columnas_interes, how='all')

            if not df_seleccionado.empty:
                df_seleccionado['Archivo'] = os.path.splitext(archivo)[0] # Nombre del archivo sin extensión
                # Reordenar para consistencia antes de concatenar
                columnas_df = ['Archivo'] + [col for col in columnas_interes if col in df_seleccionado.columns] # Columnas que realmente tiene + Archivo
                df_consolidado = pd.concat([df_consolidado, df_seleccionado[columnas_df]], ignore_index=True)
                print(f"  Datos de {archivo} añadidos.")
            else:
                 print(f"  No se encontraron datos válidos en las columnas de interés en {archivo} después de limpiar NaNs.")

        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante consolidación: {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(error_msg)

    if errores:
        messagebox.showwarning("Advertencia Consolidación", "Ocurrieron errores al consolidar algunos archivos:\n\n" + "\n".join(errores))
        # Continuar de todos modos si se consolidó algo

    if df_consolidado.empty:
        print("No se consolidaron datos.")
        messagebox.showinfo("Consolidación", "No se pudo consolidar ningún dato de los archivos encontrados.")
        return None

    # Asegurar orden final y presencia de todas las columnas de interés + Archivo
    columnas_finales_ordenadas = ['Archivo'] + columnas_interes
    # Añadir columnas faltantes en el DF final si alguna nunca apareció
    for col in columnas_finales_ordenadas:
        if col not in df_consolidado.columns:
            df_consolidado[col] = pd.NA
    # Reordenar al orden deseado
    df_consolidado = df_consolidado[columnas_finales_ordenadas]

    try:
        df_consolidado.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"Archivo consolidado creado/actualizado: {ruta_salida}")
        return ruta_salida # Devolver ruta del archivo creado
    except Exception as e:
        error_msg = f"Error al guardar el archivo consolidado '{ruta_salida}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Consolidación", f"No se pudo guardar el archivo consolidado:\n{error_msg}")
        return None

def _asignar_materiales_interno(ruta_archivo_consolidado):
    """
    Asigna materiales basándose en la columna 'Espesor' del archivo consolidado.
    Sobrescribe el archivo consolidado con la nueva columna 'Material'.
    Retorna True si tiene éxito, False si falla.
    """
    print(f"--- Iniciando asignación de materiales en: {ruta_archivo_consolidado} ---")
    if not ruta_archivo_consolidado or not os.path.exists(ruta_archivo_consolidado):
        messagebox.showerror("Error Asignación", f"El archivo consolidado '{ruta_archivo_consolidado}' no existe o no se pudo crear en el paso anterior.")
        return False

    # Mapa de espesores basado en el código original (programa_v1.1.py)
    mapa_espesores = {
        16: 'MELAMINA 16 MM',
        19: 'EUCALIPTO 18 MM', # Original tenía 19 -> Eucalipto 18mm
        20: 'MADERA MELINA',
        18: 'MELAMINA 18 MM',
        22: 'MADERA MELINA'
        # Añadir otros mapeos si son necesarios, ej: 17?
        # 17: 'MELAMINA 18 MM' # Ejemplo si 17 también va a melamina 18
    }

    try:
        df = pd.read_excel(ruta_archivo_consolidado, engine='openpyxl')

        if 'Espesor' not in df.columns:
            messagebox.showerror("Error Asignación", "La columna 'Espesor' no se encuentra en el archivo consolidado.")
            return False

        # Función para convertir 'Espesor' a un tipo comparable (int si es posible)
        def convertir_a_int_o_mantener(val):
            if pd.isna(val):
                return None
            try:
                # Intentar convertir a numérico y luego a entero
                num = pd.to_numeric(val)
                if pd.notna(num) and num == int(num):
                    return int(num)
                return val # Devolver original si no es entero (o ya era string)
            except:
                return val # Devolver original en caso de error

        # Aplicar conversión y luego mapeo
        df['Espesor_Convertido'] = df['Espesor'].apply(convertir_a_int_o_mantener)
        df['Material'] = df['Espesor_Convertido'].map(mapa_espesores).fillna('Material no definido') # Default si no hay mapeo

        # Eliminar columna temporal de conversión
        df = df.drop(columns=['Espesor_Convertido'])

        # Reordenar columnas para poner 'Material' después de 'Espesor'
        cols = df.columns.tolist()
        if 'Material' in cols:
            cols.remove('Material')
        if 'Espesor' in cols:
            idx_espesor = cols.index('Espesor')
            cols.insert(idx_espesor + 1, 'Material')
        else:
            cols.append('Material') # Añadir al final si 'Espesor' no estaba
        df = df[cols]

        # Sobrescribir el archivo consolidado
        df.to_excel(ruta_archivo_consolidado, index=False, engine='openpyxl')
        print(f"Columna 'Material' añadida y archivo guardado: {ruta_archivo_consolidado}")
        return True

    except Exception as e:
        error_msg = f"Error durante la asignación de materiales en '{ruta_archivo_consolidado}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Asignación", f"Error al asignar materiales:\n{error_msg}")
        return False

# --- Funciones de los Botones Principales ---

def procesar_y_asignar_flujo_completo():
    """
    BOTÓN 1: Ejecuta el flujo completo de procesamiento:
    1. Seleccionar Directorio.
    2. Eliminar 'mm' (interno).
    3. Consolidar Datos (interno) -> Fusion.xlsx.
    4. Asignar Materiales (interno) en Fusion.xlsx.
    """
    directorio = seleccionar_carpeta("Selecciona el directorio con los archivos Excel a procesar")
    if not directorio:
        messagebox.showwarning("Cancelado", "No se seleccionó ningún directorio.")
        return

    nombre_archivo_fusion = "Fusion.xlsx" # Nombre fijo para el archivo consolidado

    # Paso 1: Eliminar 'mm'
    print("\n--- PASO 1: Eliminando 'mm' ---")
    if not _eliminar_mm_interno(directorio, nombre_archivo_fusion):
        # Mensaje de error ya mostrado si falla gravemente
        print("Proceso detenido debido a errores durante la eliminación de 'mm'.")
        # Podríamos preguntar si continuar, pero por ahora detenemos.
        return

    # Paso 2: Consolidar Datos
    print("\n--- PASO 2: Consolidando datos ---")
    ruta_consolidado = _consolidar_datos_interno(directorio, nombre_archivo_fusion)
    if not ruta_consolidado:
        # Mensaje de error/info ya mostrado si falla o no hay datos
        print(f"Proceso detenido. No se pudo crear o no hay datos para '{nombre_archivo_fusion}'.")
        return

    # Paso 3: Asignar Materiales
    print("\n--- PASO 3: Asignando materiales ---")
    if _asignar_materiales_interno(ruta_consolidado):
        messagebox.showinfo("Éxito", f"Proceso completado.\n\nSe eliminó 'mm', se consolidaron los datos y se asignaron materiales en:\n{ruta_consolidado}")
    else:
        # Mensaje de error ya mostrado
        print(f"El proceso de asignación de materiales falló. El archivo {ruta_consolidado} puede estar incompleto.")
        messagebox.showerror("Error Final", f"El proceso de asignación de materiales falló.\nRevise el archivo {ruta_consolidado} y los logs de consola.")

def insertar_datos_en_hoja_destino():
    """
    BOTÓN 2: Inserta los datos del archivo consolidado (Fusion.xlsx)
    en la hoja 'LISTA' de un archivo Excel destino seleccionado.
    """
    print("\n--- Iniciando inserción en hoja destino ---")
    # 1. Seleccionar archivo fuente (el consolidado)
    nombre_archivo_esperado = "Fusion.xlsx"
    source_file = seleccionar_archivo(
        f"Seleccione el archivo CONSOLIDADO ({nombre_archivo_esperado})",
        # Permitir seleccionar otros .xlsx pero advertir si no es el esperado
        filetypes=[("Archivo Consolidado", nombre_archivo_esperado), ("Archivos Excel", "*.xlsx")]
    )
    if not source_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo fuente consolidado.")
        return

    # Advertencia si el nombre no es el esperado
    if os.path.basename(source_file) != nombre_archivo_esperado:
        if not messagebox.askyesno("Confirmar Fuente", f"El archivo seleccionado no se llama '{nombre_archivo_esperado}'.\n({os.path.basename(source_file)})\n\n¿Está seguro que es el archivo procesado por el Paso 1 (con materiales asignados)?"):
            return

    # 2. Leer archivo fuente
    try:
        print(f"Leyendo archivo fuente: {source_file}")
        source_data = pd.read_excel(source_file, engine='openpyxl')
        # Verificar columnas necesarias para la inserción (basado en el código original de inserción)
        required_cols = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor', 'Material', 'Archivo']
        missing_cols = [col for col in required_cols if col not in source_data.columns]
        if missing_cols:
            messagebox.showerror("Error Columnas", f"El archivo fuente '{os.path.basename(source_file)}' no contiene las columnas necesarias: {', '.join(missing_cols)}")
            return
        print("Columnas requeridas encontradas en el archivo fuente.")
    except Exception as e:
        messagebox.showerror("Error Lectura Fuente", f"Error al leer el archivo fuente '{os.path.basename(source_file)}':\n{e}\n{traceback.format_exc()}")
        return

    # 3. Seleccionar archivo destino
    destination_file = seleccionar_archivo("Seleccione el archivo Excel DESTINO (con la hoja 'LISTA')")
    if not destination_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo destino.")
        return

    # 4. Cargar archivo destino y hoja 'LISTA'
    sheet_name = 'LISTA' # Nombre de hoja fijo
    try:
        print(f"Cargando archivo destino: {destination_file}")
        wb = load_workbook(destination_file)
        if sheet_name not in wb.sheetnames:
             # Intentar con 'lista' en minúscula por si acaso
             sheet_name_lower = sheet_name.lower()
             if sheet_name_lower in (s.lower() for s in wb.sheetnames):
                 # Encontrar el nombre exacto con el case correcto
                 sheet_name = next(s for s in wb.sheetnames if s.lower() == sheet_name_lower)
                 print(f"Advertencia: Se usará la hoja '{sheet_name}' (encontrada como '{sheet_name_lower}').")
             else:
                 messagebox.showerror("Error Hoja Destino", f"La hoja '{sheet_name}' (o '{sheet_name_lower}') no se encuentra en el archivo destino.")
                 return
        ws = wb[sheet_name]
        print(f"Hoja '{sheet_name}' encontrada en el archivo destino.")
    except Exception as e:
        messagebox.showerror("Error Carga Destino", f"Error al cargar el archivo destino '{os.path.basename(destination_file)}':\n{e}\n{traceback.format_exc()}")
        return

    # 5. Insertar datos
    start_row = 13 # Fila inicial (1-based index)
    print(f"Iniciando inserción de datos en la hoja '{sheet_name}' a partir de la fila {start_row}...")

    # Mapeo de materiales a códigos (del código original)
    codigo_map = {
        'MELAMINA 18 MM': 2,
        'MELAMINA 16 MM': 2,
        'EUCALIPTO 18 MM': 2,
        'MADERA MELINA': 1
        # Añadir otros si es necesario, o un default
    }
    # Colores (del código original)
    color_amarillo = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    color_azul = styles.PatternFill(start_color="40A2E3", end_color="40A2E3", fill_type="solid")
    no_color = styles.PatternFill(fill_type=None) # Para quitar color si es necesario

    # Opcional: Limpiar filas antiguas antes de insertar?
    # max_row_actual = ws.max_row
    # if max_row_actual >= start_row:
    #    print(f"Limpiando filas existentes desde {start_row} hasta {max_row_actual}")
    #    ws.delete_rows(start_row, max_row_actual - start_row + 1)

    try:
        for index, row_data in source_data.iterrows():
            row_num = start_row + index
            material = row_data.get('Material', '') # Usar .get para seguridad
            codigo = codigo_map.get(material, '') # Obtener código, default vacío ''

            # Escribir datos en las celdas correspondientes
            ws[f'A{row_num}'] = row_data.get('QTY', '')
            ws[f'B{row_num}'] = row_data.get('DESCRIPTION', '')
            ws[f'C{row_num}'] = row_data.get('Largo', '')
            ws[f'D{row_num}'] = row_data.get('Ancho', '')
            ws[f'E{row_num}'] = row_data.get('Espesor', '')
            ws[f'F{row_num}'] = codigo # Columna CODIGO
            ws[f'G{row_num}'] = material # Columna MATERIAL
            ws[f'H{row_num}'] = row_data.get('Archivo', '') # Columna Archivo Origen

            # Aplicar formato de color a la celda del código (Columna F)
            cell_f = ws[f'F{row_num}']
            if codigo == 1:
                cell_f.fill = color_amarillo
            elif codigo == 2:
                cell_f.fill = color_azul
            else:
                cell_f.fill = no_color # Quitar color si no hay código 1 o 2

            if (index + 1) % 50 == 0: # Log de progreso
                print(f"  Insertadas {index + 1} filas...")

        print(f"Inserción completada. Total de {len(source_data)} filas procesadas.")

        # 6. Guardar el archivo destino
        wb.save(destination_file)
        print(f"Archivo destino guardado: {destination_file}")
        messagebox.showinfo("Éxito", f"Datos insertados y formateados correctamente desde\n'{os.path.basename(source_file)}'\nen la hoja '{sheet_name}' del archivo\n'{os.path.basename(destination_file)}'.")

    except Exception as e:
        error_msg = f"Error durante la inserción de datos o al guardar el archivo destino:\n{e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Inserción/Guardado", error_msg)


# --- Configuración de la Interfaz Gráfica Principal ---
ventana_principal = tk.Tk()
ventana_principal.title("Listas COPT v2.0")
ventana_principal.geometry("450x300")  # Ajustar tamaño para dos botones + info
ventana_principal.configure(bg="#F8EDE3")  # Color de fondo original

# Estilo de botones (del código original)
style = ttk.Style()
style.configure("TButton",
                padding=10,
                relief="flat",
                background="#D0B8A8",
                foreground="#E36414",
                font=("Times New Roman", 13))
style.map("TButton",
          background=[('active', '#D0B8A8')], # Mantener color en hover/active
          foreground=[('active', '#9A031E')])

# Crear los botones principales
label1 = ttk.Label(ventana_principal, text="Paso 1: Procesar archivos y generar 'Fusion.xlsx'", background="#F8EDE3", font=('Times New Roman', 11))
label1.pack(pady=(15, 2))
boton_procesar = ttk.Button(ventana_principal,
                            text="Procesar Carpeta y Asignar Materiales",
                            command=procesar_y_asignar_flujo_completo,
                            width=35)
boton_procesar.pack(pady=5, padx=20)

label2 = ttk.Label(ventana_principal, text="Paso 2: Cargar 'Fusion.xlsx' en la hoja 'LISTA' destino", background="#F8EDE3", font=('Times New Roman', 11))
label2.pack(pady=(15, 2))
boton_insertar = ttk.Button(ventana_principal,
                            text="Insertar Datos en Hoja Destino",
                            command=insertar_datos_en_hoja_destino,
                            width=35)
boton_insertar.pack(pady=5, padx=20)

# Info Desarrollador y Versión (del código original)
def abrir_ayuda():
    webbrowser.open("https://www.pentestercr.com/") # Mantener enlace original

ttk.Button(ventana_principal, text="Desarrollado por S4mma3l", command=abrir_ayuda, style="TButton").pack(pady=(20, 5), padx=40) # Aplicar estilo
enlace = tk.Label(text="Version: 2.0.0", font=("Times New Roman", 10,), fg="#5F0F40", bg="#F8EDE3")
enlace.pack(pady=0.5)

# Iniciar el bucle principal de la GUI
ventana_principal.mainloop()

# --- END OF FILE programa_v2.0.py ---