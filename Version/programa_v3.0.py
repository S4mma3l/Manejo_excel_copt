# --- START OF FILE programa_v2.1_moderno.py ---

import pandas as pd
import os
from openpyxl import load_workbook, styles
# Import tkinter components that ttkbootstrap doesn't replace
from tkinter import filedialog, messagebox
# Import ttkbootstrap instead of tkinter.ttk and optionally tk
import ttkbootstrap as ttk
from ttkbootstrap.constants import * # For constants like PRIMARY, INFO, etc.
import webbrowser
import traceback

# --- Funciones Auxiliares (sin cambios) ---

def seleccionar_carpeta(titulo):
    """Función para seleccionar una carpeta."""
    # No se necesita la ventana raíz explícita aquí, filedialog la maneja
    carpeta = filedialog.askdirectory(title=titulo)
    if not carpeta:
        return None
    return carpeta

def seleccionar_archivo(titulo, filetypes=[("Archivos Excel", "*.xlsx")]):
    """Función para seleccionar un archivo."""
    archivo = filedialog.askopenfilename(title=titulo, filetypes=filetypes)
    if not archivo:
        return None
    return archivo

# --- Funciones Internas del Flujo Principal (sin cambios en lógica) ---
# (Se mantienen _eliminar_mm_interno, _consolidar_datos_interno, _asignar_materiales_interno
# como en la versión anterior v2.0)

def _eliminar_mm_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    """
    Elimina 'mm' de las celdas de texto en archivos Excel de un directorio.
    Retorna True si se completa (incluso si no hay archivos), False si hay errores graves.
    """
    print(f"--- Iniciando eliminación de 'mm' en: {directorio} ---")
    try:
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}", parent=ventana_principal)
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}", parent=ventana_principal)
        return False

    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para procesar.")
        return True

    errores = []
    modificado_alguno = False
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para eliminar 'mm'...")
            df_dict = pd.read_excel(ruta_archivo, sheet_name=None, engine='openpyxl') # Lee todas las hojas en un dict
            df_modificados = {}
            archivo_modificado_en_hoja = False

            for sheet_name, df_hoja in df_dict.items():
                # Convertir todo a string temporalmente para reemplazo seguro? NO, applymap maneja tipos
                # df_str = df_hoja.astype(str) # No es estrictamente necesario con applymap si la lambda maneja tipos

                # --- CORRECCIÓN AQUÍ ---
                # Usar applymap para aplicar la función a cada celda del DataFrame
                df_hoja_mod = df_hoja.applymap(lambda x: x.replace('mm', '').replace('MM', '').strip() if isinstance(x, str) else x)
                # --- FIN CORRECCIÓN ---

                # Verificar si hubo cambios reales antes de marcar como modificado
                # (Comparar DataFrames puede ser costoso, pero es más preciso)
                # Una forma más simple es verificar si la función *podría* haber hecho un cambio
                # if df_str.apply(lambda col: col.str.contains('mm', case=False, na=False)).any().any(): # Verificación previa (opcional)

                if not df_hoja.equals(df_hoja_mod):
                    df_modificados[sheet_name] = df_hoja_mod
                    archivo_modificado_en_hoja = True
                    print(f"  'mm' encontrado y eliminado en hoja '{sheet_name}' de {archivo}.")
                else:
                    df_modificados[sheet_name] = df_hoja # Mantener la hoja original si no hubo cambios reales

            if archivo_modificado_en_hoja:
                 with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                     for sheet_name, df_final in df_modificados.items():
                         df_final.to_excel(writer, sheet_name=sheet_name, index=False)
                 modificado_alguno = True
                 print(f"  Archivo '{archivo}' modificado y guardado.")
            else:
                print(f"  No se realizaron modificaciones detectables en {archivo}.")

        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante eliminación de 'mm': {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(f"Error procesando {archivo}: {e}") # Mensaje más corto para el messagebox

    if errores:
        messagebox.showwarning("Advertencia Eliminación 'mm'", "Ocurrieron errores al eliminar 'mm' en algunos archivos:\n\n" + "\n".join(errores), parent=ventana_principal)
    elif modificado_alguno:
         print("Proceso de eliminación de 'mm' completado con modificaciones.")
    else:
         print("Proceso de eliminación de 'mm' completado. No se realizaron modificaciones.")

    print("--- Finalizada eliminación de 'mm' ---")
    return True

def _consolidar_datos_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    """
    Consolida datos de archivos Excel en un directorio especificado.
    Retorna la ruta del archivo consolidado si tiene éxito y datos, None si falla o no hay datos.
    (Lógica idéntica a la versión anterior)
    """
    print(f"--- Iniciando consolidación de datos en: {directorio} ---")
    columnas_interes = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor']
    df_consolidado = pd.DataFrame()
    ruta_salida = os.path.join(directorio, archivo_consolidado_nombre)

    try:
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}", parent=ventana_principal) # Añadir parent
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}", parent=ventana_principal) # Añadir parent
        return None

    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para consolidar.")
        messagebox.showinfo("Consolidación", "No se encontraron archivos Excel (aparte de {} si existe) para consolidar.".format(archivo_consolidado_nombre), parent=ventana_principal) # Añadir parent
        return None

    errores = []
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para consolidación...")
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            columnas_disponibles = [col for col in columnas_interes if col in df.columns]
            if not columnas_disponibles:
                print(f"  Advertencia: Ninguna de las columnas de interés {columnas_interes} encontrada en {archivo}. Saltando archivo.")
                continue

            df_seleccionado = df[columnas_disponibles].copy()
            for col in columnas_interes:
                if col not in df_seleccionado.columns:
                    df_seleccionado[col] = pd.NA
            df_seleccionado = df_seleccionado.dropna(subset=columnas_interes, how='all')

            if not df_seleccionado.empty:
                df_seleccionado['Archivo'] = os.path.splitext(archivo)[0]
                columnas_df = ['Archivo'] + [col for col in columnas_interes if col in df_seleccionado.columns]
                df_consolidado = pd.concat([df_consolidado, df_seleccionado[columnas_df]], ignore_index=True)
                print(f"  Datos de {archivo} añadidos.")
            else:
                 print(f"  No se encontraron datos válidos en las columnas de interés en {archivo} después de limpiar NaNs.")

        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante consolidación: {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(error_msg)

    if errores:
        messagebox.showwarning("Advertencia Consolidación", "Ocurrieron errores al consolidar algunos archivos:\n\n" + "\n".join(errores), parent=ventana_principal) # Añadir parent

    if df_consolidado.empty:
        print("No se consolidaron datos.")
        messagebox.showinfo("Consolidación", "No se pudo consolidar ningún dato de los archivos encontrados.", parent=ventana_principal) # Añadir parent
        return None

    columnas_finales_ordenadas = ['Archivo'] + columnas_interes
    for col in columnas_finales_ordenadas:
        if col not in df_consolidado.columns:
            df_consolidado[col] = pd.NA
    df_consolidado = df_consolidado[columnas_finales_ordenadas]

    try:
        df_consolidado.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"Archivo consolidado creado/actualizado: {ruta_salida}")
        return ruta_salida
    except Exception as e:
        error_msg = f"Error al guardar el archivo consolidado '{ruta_salida}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Consolidación", f"No se pudo guardar el archivo consolidado:\n{error_msg}", parent=ventana_principal) # Añadir parent
        return None

def _asignar_materiales_interno(ruta_archivo_consolidado):
    """
    Asigna materiales basándose en la columna 'Espesor' del archivo consolidado.
    Sobrescribe el archivo consolidado con la nueva columna 'Material'.
    Retorna True si tiene éxito, False si falla.
    (Lógica idéntica a la versión anterior)
    """
    print(f"--- Iniciando asignación de materiales en: {ruta_archivo_consolidado} ---")
    if not ruta_archivo_consolidado or not os.path.exists(ruta_archivo_consolidado):
        messagebox.showerror("Error Asignación", f"El archivo consolidado '{os.path.basename(ruta_archivo_consolidado)}' no existe o no se pudo crear.", parent=ventana_principal) # Añadir parent
        return False

    mapa_espesores = {
        16: 'MELAMINA 16 MM', 19: 'EUCALIPTO 18 MM', 20: 'MADERA MELINA',
        18: 'MELAMINA 18 MM', 22: 'MADERA MELINA', 10: 'MADERA MELINA',
        25: 'BUTCHERBLOCK', 3: 'MADERA MELINA', 2: 'MADERA MELINA', 42: 'BUTCHERBLOCK', 50: 'BUTCHERBLOCK',
    }

    try:
        df = pd.read_excel(ruta_archivo_consolidado, engine='openpyxl')
        if 'Espesor' not in df.columns:
            messagebox.showerror("Error Asignación", "La columna 'Espesor' no se encuentra en el archivo consolidado.", parent=ventana_principal) # Añadir parent
            return False

        def convertir_a_int_o_mantener(val):
            if pd.isna(val): return None
            try:
                num = pd.to_numeric(val)
                if pd.notna(num) and num == int(num): return int(num)
                return val
            except: return val

        df['Espesor_Convertido'] = df['Espesor'].apply(convertir_a_int_o_mantener)
        df['Material'] = df['Espesor_Convertido'].map(mapa_espesores).fillna('Material no definido')
        df = df.drop(columns=['Espesor_Convertido'])

        cols = df.columns.tolist()
        if 'Material' in cols: cols.remove('Material')
        if 'Espesor' in cols:
            idx_espesor = cols.index('Espesor')
            cols.insert(idx_espesor + 1, 'Material')
        else: cols.append('Material')
        df = df[cols]

        df.to_excel(ruta_archivo_consolidado, index=False, engine='openpyxl')
        print(f"Columna 'Material' añadida y archivo guardado: {ruta_archivo_consolidado}")
        return True

    except Exception as e:
        error_msg = f"Error durante la asignación de materiales en '{os.path.basename(ruta_archivo_consolidado)}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Asignación", f"Error al asignar materiales:\n{error_msg}", parent=ventana_principal) # Añadir parent
        return False


# --- Funciones de los Botones Principales (sin cambios en lógica) ---

def procesar_y_asignar_flujo_completo():
    """
    BOTÓN 1: Ejecuta el flujo completo de procesamiento.
    (Lógica idéntica a la versión anterior)
    """
    directorio = seleccionar_carpeta("Selecciona el directorio con los archivos Excel a procesar")
    if not directorio:
        messagebox.showwarning("Cancelado", "No se seleccionó ningún directorio.", parent=ventana_principal) # Añadir parent
        return

    nombre_archivo_fusion = "Fusion.xlsx"

    print("\n--- PASO 1: Eliminando 'mm' ---")
    if not _eliminar_mm_interno(directorio, nombre_archivo_fusion):
        print("Proceso detenido debido a errores durante la eliminación de 'mm'.")
        return

    print("\n--- PASO 2: Consolidando datos ---")
    ruta_consolidado = _consolidar_datos_interno(directorio, nombre_archivo_fusion)
    if not ruta_consolidado:
        print(f"Proceso detenido. No se pudo crear o no hay datos para '{nombre_archivo_fusion}'.")
        return

    print("\n--- PASO 3: Asignando materiales ---")
    if _asignar_materiales_interno(ruta_consolidado):
        messagebox.showinfo("Éxito", f"Proceso completado.\n\nSe eliminó 'mm', se consolidaron los datos y se asignaron materiales en:\n{ruta_consolidado}", parent=ventana_principal) # Añadir parent
    else:
        print(f"El proceso de asignación de materiales falló. El archivo {ruta_consolidado} puede estar incompleto.")
        messagebox.showerror("Error Final", f"El proceso de asignación de materiales falló.\nRevise el archivo {ruta_consolidado} y los logs de consola.", parent=ventana_principal) # Añadir parent

def insertar_datos_en_hoja_destino():
    """
    BOTÓN 2: Inserta los datos del archivo consolidado (Fusion.xlsx)
    en la hoja 'LISTA' de un archivo Excel destino seleccionado.
    (Lógica idéntica a la versión anterior)
    """
    print("\n--- Iniciando inserción en hoja destino ---")
    nombre_archivo_esperado = "Fusion.xlsx"
    source_file = seleccionar_archivo(
        f"Seleccione el archivo CONSOLIDADO ({nombre_archivo_esperado})",
        filetypes=[("Archivo Consolidado", nombre_archivo_esperado), ("Archivos Excel", "*.xlsx")]
    )
    if not source_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo fuente.", parent=ventana_principal) # Añadir parent
        return

    if os.path.basename(source_file) != nombre_archivo_esperado:
        if not messagebox.askyesno("Confirmar Fuente", f"El archivo seleccionado no se llama '{nombre_archivo_esperado}'.\n({os.path.basename(source_file)})\n\n¿Continuar usando este archivo?", parent=ventana_principal): # Añadir parent
            return

    try:
        print(f"Leyendo archivo fuente: {source_file}")
        source_data = pd.read_excel(source_file, engine='openpyxl')
        required_cols = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor', 'Material', 'Archivo']
        missing_cols = [col for col in required_cols if col not in source_data.columns]
        if missing_cols:
            messagebox.showerror("Error Columnas", f"El archivo fuente '{os.path.basename(source_file)}' no contiene las columnas necesarias: {', '.join(missing_cols)}", parent=ventana_principal) # Añadir parent
            return
        print("Columnas requeridas encontradas.")
    except Exception as e:
        messagebox.showerror("Error Lectura Fuente", f"Error al leer el archivo fuente:\n{e}\n{traceback.format_exc()}", parent=ventana_principal) # Añadir parent
        return

    destination_file = seleccionar_archivo("Seleccione el archivo Excel DESTINO (con la hoja 'LISTA')")
    if not destination_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo destino.", parent=ventana_principal) # Añadir parent
        return

    sheet_name = 'LISTA'
    try:
        print(f"Cargando archivo destino: {destination_file}")
        wb = load_workbook(destination_file)
        sheet_name_found = None
        for s_name in wb.sheetnames:
            if s_name.upper() == sheet_name.upper(): # Case-insensitive check
                sheet_name_found = s_name
                break
        if not sheet_name_found:
             messagebox.showerror("Error Hoja Destino", f"La hoja '{sheet_name}' no se encuentra en el archivo destino.", parent=ventana_principal) # Añadir parent
             return
        ws = wb[sheet_name_found]
        print(f"Hoja '{sheet_name_found}' encontrada en el archivo destino.")
    except Exception as e:
        messagebox.showerror("Error Carga Destino", f"Error al cargar el archivo destino:\n{e}\n{traceback.format_exc()}", parent=ventana_principal) # Añadir parent
        return

    start_row = 13
    print(f"Iniciando inserción de datos en la hoja '{sheet_name_found}' a partir de la fila {start_row}...")
    codigo_map = {'MELAMINA 18 MM': 2, 'MELAMINA 16 MM': 2, 'EUCALIPTO 18 MM': 2, 'MADERA MELINA': 1, 'BUTCHERBLOCK': 1}
    # Definición de colores para las celdas
    color_amarillo = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Amarillo
    color_azul = styles.PatternFill(start_color="40A2E3", end_color="40A2E3", fill_type="solid") # Azul
    no_color = styles.PatternFill(fill_type=None) # Sin color

    try:
        for index, row_data in source_data.iterrows():
            row_num = start_row + index
            material = row_data.get('Material', '')
            codigo = codigo_map.get(material, '')

            ws[f'A{row_num}'] = row_data.get('QTY', '')
            ws[f'B{row_num}'] = row_data.get('DESCRIPTION', '')
            ws[f'C{row_num}'] = row_data.get('Largo', '')
            ws[f'D{row_num}'] = row_data.get('Ancho', '')
            ws[f'E{row_num}'] = row_data.get('Espesor', '')
            ws[f'F{row_num}'] = codigo
            ws[f'G{row_num}'] = material
            ws[f'H{row_num}'] = row_data.get('Archivo', '')

            cell_f = ws[f'F{row_num}']
            if codigo == 1: cell_f.fill = color_amarillo
            elif codigo == 2: cell_f.fill = color_azul
            else: cell_f.fill = no_color

            if (index + 1) % 100 == 0: print(f"  Insertadas {index + 1} filas...") # Log every 100 rows

        print(f"Inserción completada. Total de {len(source_data)} filas procesadas.")

        wb.save(destination_file)
        print(f"Archivo destino guardado: {destination_file}")
        messagebox.showinfo("Éxito", f"Datos insertados y formateados correctamente desde\n'{os.path.basename(source_file)}'\nen la hoja '{sheet_name_found}' del archivo\n'{os.path.basename(destination_file)}'.", parent=ventana_principal) # Añadir parent

    except Exception as e:
        error_msg = f"Error durante la inserción o guardado:\n{e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Inserción/Guardado", error_msg, parent=ventana_principal) # Añadir parent


# --- Configuración de la Interfaz Gráfica Principal con ttkbootstrap ---

# 1. Crear la ventana principal usando ttk.Window
# Elige un tema: cosmo, flatly, journal, litera, lumen, minty, pulse, sandstone,
# united, yeti (claros) / cyborg, darkly, solar, superhero (oscuros)
ventana_principal = ttk.Window(themename="cyborg") # <- Elige tu tema preferido aquí
ventana_principal.title("Gestor de Listas COPT v2.1")
ventana_principal.geometry("550x350") # Ajustar tamaño si es necesario

# 2. Crear un Frame principal para mejor organización y padding
main_frame = ttk.Frame(ventana_principal, padding=(20, 10)) # Padding (horizontal, vertical)
main_frame.pack(fill=BOTH, expand=YES)

# 3. Añadir Widgets usando ttk de ttkbootstrap y 'bootstyle'
# Botón 1: Procesar
label1 = ttk.Label(main_frame, text="Paso 1: Procesar archivos y generar 'Fusion.xlsx'", font="-weight bold") # Texto en negrita
label1.pack(pady=(10, 5))
boton_procesar = ttk.Button(main_frame,
                            text="Procesar Carpeta y Asignar Materiales",
                            command=procesar_y_asignar_flujo_completo,
                            bootstyle=PRIMARY) # Estilo: PRIMARY, SUCCESS, INFO, WARNING, DANGER, SECONDARY, LIGHT, DARK
boton_procesar.pack(pady=5, fill=X, padx=10) # fill=X para que ocupe el ancho

# Separador visual (opcional)
separator1 = ttk.Separator(main_frame, orient=HORIZONTAL)
separator1.pack(pady=15, fill=X, padx=10)

# Botón 2: Insertar
label2 = ttk.Label(main_frame, text="Paso 2: Cargar 'Fusion.xlsx' en la hoja 'LISTA' destino", font="-weight bold")
label2.pack(pady=(5, 5))
boton_insertar = ttk.Button(main_frame,
                            text="Insertar Datos en Hoja Destino",
                            command=insertar_datos_en_hoja_destino,
                            bootstyle=INFO) # Otro estilo para diferenciar
boton_insertar.pack(pady=5, fill=X, padx=10)

# Separador visual (opcional)
separator2 = ttk.Separator(main_frame, orient=HORIZONTAL)
separator2.pack(pady=15, fill=X, padx=10)

# Info Desarrollador y Versión
def abrir_ayuda():
    webbrowser.open("https://www.pentestercr.com/")

# Usar bootstyle="link" para el botón de desarrollador
boton_desarrollador = ttk.Button(main_frame,
                                 text="Desarrollado por S4mma3l",
                                 command=abrir_ayuda,
                                 bootstyle=(LINK, SECONDARY)) # Estilo tipo enlace
boton_desarrollador.pack(pady=(10, 0))

version_label = ttk.Label(main_frame, text="Version: 2.1.0", bootstyle=SECONDARY) # Estilo secundario para la versión
version_label.pack(pady=(0, 10))


# 4. Añadir parent=ventana_principal a todos los messagebox para que se muestren sobre la ventana principal
# (Hecho en las funciones de arriba)

# Iniciar el bucle principal de la GUI
ventana_principal.mainloop()

# --- END OF FILE programa_v2.1_moderno.py ---