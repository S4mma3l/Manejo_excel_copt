import pandas as pd
import os
from openpyxl import load_workbook, styles
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import webbrowser

# Funciones
def seleccionar_carpeta(titulo):
    """Función para seleccionar una carpeta usando un diálogo de selección de carpetas"""
    carpeta = filedialog.askdirectory(title=titulo)
    if not carpeta:
        messagebox.showerror("Error", "No se seleccionó ninguna carpeta.")
        return None
    return carpeta

def seleccionar_archivo(titulo):
    """Función para seleccionar un archivo usando un diálogo de selección de archivos"""
    archivo = filedialog.askopenfilename(title=titulo, filetypes=[("Archivos Excel", "*.xlsx")])
    if not archivo:
        messagebox.showerror("Error", "No se seleccionó ningún archivo.")
        return None
    return archivo

def eliminar_mm():
    directorio = seleccionar_carpeta("Selecciona el directorio con archivos Excel")
    if not directorio:
        return

    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            df = df.applymap(lambda x: x.replace('mm', '') if isinstance(x, str) else x)
            df.to_excel(ruta_archivo, index=False, engine='openpyxl')
            print(f"Se han eliminado las letras 'mm' de {archivo}.")
        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

def consolidar_datos():
    directorio = seleccionar_carpeta("Selecciona el directorio con archivos Excel")
    if not directorio:
        return

    columnas_interes = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor']
    df_consolidado = pd.DataFrame()
    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]

    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            columnas_disponibles = [col for col in columnas_interes if col in df.columns]
            df_seleccionado = df[columnas_disponibles].dropna(how='all')
            df_seleccionado['Archivo'] = os.path.splitext(archivo)[0]
            df_consolidado = pd.concat([df_consolidado, df_seleccionado], ignore_index=True)
        except Exception as e:
            print(f"Error al procesar {archivo}: {e}")

    columnas_finales = ['Archivo'] + columnas_interes
    columnas_finales = [col for col in columnas_finales if col in df_consolidado.columns]
    df_consolidado = df_consolidado[columnas_finales]
    ruta_salida = os.path.join(directorio, 'Fusion.xlsx')
    try:
        df_consolidado.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"Se ha creado el archivo consolidado: {ruta_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo consolidado: {e}")

def asignar_materiales():
    directorio = seleccionar_carpeta("Selecciona el directorio con archivos Excel")
    if not directorio:
        return

    archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx')]
    if not archivos_excel:
        messagebox.showerror("Error", "No hay archivos Excel en el directorio.")
        return

    def procesar_archivo():
        seleccion = listbox.curselection()
        if seleccion:
            opcion = seleccion[0]
            archivo_seleccionado = archivos_excel[opcion]
            ruta_archivo = os.path.join(directorio, archivo_seleccionado)
            mapa_espesores = {16: 'MELAMINA 16 MM', 19: 'EUCALIPTO 18 MM', 20: 'MADERA MELINA', 18: 'MELAMINA 18 MM', 22: 'MADERA MELINA'}
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            if 'Espesor' in df.columns:
                df['Material'] = df['Espesor'].map(mapa_espesores)
                nombre_archivo_modificado = f'Asignacion_de_Materiales_{archivo_seleccionado}'
                ruta_salida = os.path.join(directorio, nombre_archivo_modificado)
                try:
                    df.to_excel(ruta_salida, index=False, engine='openpyxl')
                    messagebox.showinfo("Éxito", f"Se ha añadido la columna 'Material' y se ha guardado el archivo modificado: {ruta_salida}")
                except Exception as e:
                    messagebox.showerror("Error", f"Error al guardar el archivo modificado: {e}")
            else:
                messagebox.showerror("Error", "La columna 'Espesor' no se encuentra en el archivo.")
        else:
            messagebox.showerror("Error", "No se seleccionó ningún archivo.")

    ventana_asignar = tk.Toplevel()
    ventana_asignar.title("Asignar Materiales")
    ventana_asignar.geometry("600x600")  # Ajustar tamaño de la ventana

    ttk.Label(ventana_asignar, text="Selecciona el Archivo Fusion.xlsx:").pack(pady=5)

    listbox_frame = ttk.Frame(ventana_asignar)
    listbox_frame.pack(pady=5, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL)
    listbox = tk.Listbox(listbox_frame, selectmode=tk.SINGLE, yscrollcommand=scrollbar.set, height=20)
    scrollbar.config(command=listbox.yview)

    for archivo in archivos_excel:
        listbox.insert(tk.END, archivo)

    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    ttk.Button(ventana_asignar, text="Asignar Materiales", command=procesar_archivo).pack(pady=5)

def insertar_datos_en_hoja():
    source_file = seleccionar_archivo("Seleccione el archivo fuente")
    if not source_file:
        return

    try:
        source_data = pd.read_excel(source_file, engine='openpyxl')
    except Exception as e:
        messagebox.showerror("Error", f"Error al leer el archivo fuente: {e}")
        return

    destination_file = seleccionar_archivo("Seleccione el archivo destino para cargar los datos")
    if not destination_file:
        return

    try:
        wb = load_workbook(destination_file)
        if 'LISTA' not in wb.sheetnames:
            messagebox.showerror("Error", "La hoja 'lista' no se encuentra en el archivo destino.")
            return
        ws = wb['LISTA']
    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar el archivo destino: {e}")
        return

    # Mapear los materiales a los códigos
    codigo_map = {
        'MELAMINA 18 MM': 2,
        'MELAMINA 16 MM': 2,
        'EUCALIPTO 18 MM': 2,
        'MADERA MELINA': 1
    }

    start_row = 13
    for index, row in source_data.iterrows():
        row_num = start_row + index
        material = row.get('Material', '')
        codigo = codigo_map.get(material, '')

        ws[f'A{row_num}'] = row.get('QTY', '')
        ws[f'B{row_num}'] = row.get('DESCRIPTION', '')
        ws[f'C{row_num}'] = row.get('Largo', '')
        ws[f'D{row_num}'] = row.get('Ancho', '')
        ws[f'E{row_num}'] = row.get('Espesor', '')
        ws[f'F{row_num}'] = codigo
        ws[f'G{row_num}'] = material
        ws[f'H{row_num}'] = row.get('Archivo', '')

        # Aplicar formato de color a la columna COD
        cell = ws[f'F{row_num}']
        if codigo == 1:
            cell.fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
        elif codigo == 2:
            cell.fill = styles.PatternFill(start_color="40A2E3", end_color="40A2E3", fill_type="solid")  # Azul

    try:
        wb.save(destination_file)
        messagebox.showinfo("Éxito", "Datos insertados y formateados correctamente en la hoja 'lista'.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar el archivo destino: {e}")

# Configuración de la interfaz gráfica principal
ventana_principal = tk.Tk()
ventana_principal.title("Listas COPT")
ventana_principal.geometry("320x380")  # Ajustar tamaño de la ventana principal
ventana_principal.configure(bg="#F8EDE3")  # Color de fondo

# Estilo de botones
style = ttk.Style()
style.configure("TButton",
                padding=10,
                relief="flat",
                background="#D0B8A8",  # Azul claro
                foreground="#E36414",  # Blanco
                font=("Times New Roman", 13,))
style.map("TButton",
          background=[('active', '#D0B8A8')],
          foreground=[('active', '#9A031E')])

# Agregar botones
ttk.Button(ventana_principal, text="Eliminar Milimetros(mm)", command=eliminar_mm).pack(pady=10, padx=12, fill=tk.BOTH)
ttk.Button(ventana_principal, text="Unificar Archivos Excel", command=consolidar_datos).pack(pady=10, padx=12, fill=tk.BOTH)
ttk.Button(ventana_principal, text="Asignar Materiales", command=asignar_materiales).pack(pady=10, padx=12, fill=tk.BOTH)
ttk.Button(ventana_principal, text="Insertar Datos a lista", command=insertar_datos_en_hoja).pack(pady=10, padx=12, fill=tk.BOTH)

# Añadir botón de ayuda
def abrir_ayuda():
    webbrowser.open("https://www.pentestercr.com/")

ttk.Button(ventana_principal, text="Desarrollado por S4mma3l", command=abrir_ayuda).pack(pady=10, padx=40, fill=tk.BOTH)
enlace = tk.Label(text="Version: 1.1.0", font=("Times New Roman", 10,), fg="#5F0F40", bg="#F8EDE3")
enlace.pack(pady=0.5)

ventana_principal.mainloop()
