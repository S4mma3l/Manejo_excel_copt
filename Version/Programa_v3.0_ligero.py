# --- START OF FILE programa_v2.2_ligero.py ---

# Importaciones ligeras para la UI y utilidades básicas
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font as tkFont # Añadido tkFont
import os
import webbrowser
import traceback

# --- Funciones Auxiliares (sin cambios) ---
# (seleccionar_carpeta, seleccionar_archivo)
def seleccionar_carpeta(titulo):
    # No se necesita la ventana raíz explícita aquí, filedialog la maneja
    carpeta = filedialog.askdirectory(title=titulo)
    if not carpeta:
        return None
    return carpeta

def seleccionar_archivo(titulo, filetypes=[("Archivos Excel", "*.xlsx")]):
    archivo = filedialog.askopenfilename(title=titulo, filetypes=filetypes)
    if not archivo:
        return None
    return archivo

# --- Funciones Internas del Flujo Principal (mismas optimizaciones de datos que antes) ---
# (_eliminar_mm_interno, _consolidar_datos_interno, _asignar_materiales_interno)
# Estas funciones NO cambian respecto a la versión _optimizado.py,
# ya que las optimizaciones de pandas y openpyxl son independientes de la UI.
# Solo asegúrate de que las importaciones de pandas/openpyxl están DENTRO de ellas.

def _eliminar_mm_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    import pandas as pd # IMPORTACIÓN DIFERIDA
    # ... (resto del código de _eliminar_mm_interno sin cambios) ...
    print(f"--- Iniciando eliminación de 'mm' en: {directorio} ---")
    try:
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}", parent=ventana_principal)
        return False
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}", parent=ventana_principal)
        return False

    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para procesar.")
        return True

    errores = []
    modificado_alguno_global = False
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para eliminar 'mm'...")
            df_dict = pd.read_excel(ruta_archivo, sheet_name=None, engine='openpyxl')
            df_modificados_dict = {}
            archivo_modificado_localmente = False

            for sheet_name, df_hoja_original in df_dict.items():
                modificaciones_en_hoja_actual = False
                def replace_mm_and_track(x):
                    nonlocal modificaciones_en_hoja_actual
                    if isinstance(x, str):
                        original_x = x
                        x_mod = x.replace('mm', '').replace('MM', '').strip()
                        if x_mod != original_x:
                            modificaciones_en_hoja_actual = True
                        return x_mod
                    return x
                df_hoja_procesada = df_hoja_original.applymap(replace_mm_and_track)
                if modificaciones_en_hoja_actual:
                    df_modificados_dict[sheet_name] = df_hoja_procesada
                    archivo_modificado_localmente = True
                    print(f"  'mm' encontrado y eliminado en hoja '{sheet_name}' de {archivo}.")
                else:
                    df_modificados_dict[sheet_name] = df_hoja_original
            if archivo_modificado_localmente:
                 with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                     for sheet_name_final, df_final_a_escribir in df_modificados_dict.items():
                         df_final_a_escribir.to_excel(writer, sheet_name=sheet_name_final, index=False)
                 modificado_alguno_global = True
                 print(f"  Archivo '{archivo}' modificado y guardado.")
            else:
                print(f"  No se realizaron modificaciones detectables en {archivo}.")
        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante eliminación de 'mm': {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(f"Error procesando {archivo}: {e}")
    if errores:
        messagebox.showwarning("Advertencia Eliminación 'mm'", "Ocurrieron errores al eliminar 'mm' en algunos archivos:\n\n" + "\n".join(errores), parent=ventana_principal)
    elif modificado_alguno_global:
         print("Proceso de eliminación de 'mm' completado con modificaciones.")
    else:
         print("Proceso de eliminación de 'mm' completado. No se realizaron modificaciones.")
    print("--- Finalizada eliminación de 'mm' ---")
    return True

def _consolidar_datos_interno(directorio, archivo_consolidado_nombre='Fusion.xlsx'):
    import pandas as pd # IMPORTACIÓN DIFERIDA
    # ... (resto del código de _consolidar_datos_interno sin cambios) ...
    print(f"--- Iniciando consolidación de datos en: {directorio} ---")
    columnas_interes = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor']
    lista_dfs_a_concatenar = []
    ruta_salida = os.path.join(directorio, archivo_consolidado_nombre)
    try:
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') and not f.startswith('~$') and f != archivo_consolidado_nombre]
    except FileNotFoundError:
        messagebox.showerror("Error", f"El directorio seleccionado no existe:\n{directorio}", parent=ventana_principal)
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Error al listar archivos en el directorio:\n{directorio}\n{e}", parent=ventana_principal)
        return None
    if not archivos_excel:
        print("No se encontraron archivos Excel (excluyendo temporales y de fusión) para consolidar.")
        messagebox.showinfo("Consolidación", "No se encontraron archivos Excel (aparte de {} si existe) para consolidar.".format(archivo_consolidado_nombre), parent=ventana_principal)
        return None
    errores = []
    for archivo in archivos_excel:
        ruta_archivo = os.path.join(directorio, archivo)
        try:
            print(f"Procesando {archivo} para consolidación...")
            df = pd.read_excel(ruta_archivo, engine='openpyxl')
            columnas_disponibles = [col for col in columnas_interes if col in df.columns]
            if not columnas_disponibles:
                print(f"  Advertencia: Ninguna de las columnas de interés {columnas_interes} encontrada en {archivo}. Saltando archivo.")
                continue
            df_seleccionado = df[columnas_disponibles].copy()
            for col in columnas_interes:
                if col not in df_seleccionado.columns:
                    df_seleccionado[col] = pd.NA
            df_seleccionado = df_seleccionado.dropna(subset=columnas_interes, how='all')
            if not df_seleccionado.empty:
                df_seleccionado['Archivo'] = os.path.splitext(archivo)[0]
                columnas_df = ['Archivo'] + [col for col in columnas_interes if col in df_seleccionado.columns]
                lista_dfs_a_concatenar.append(df_seleccionado[columnas_df])
                print(f"  Datos de {archivo} añadidos.")
            else:
                 print(f"  No se encontraron datos válidos en las columnas de interés en {archivo} después de limpiar NaNs.")
        except Exception as e:
            error_msg = f"Error al procesar {archivo} durante consolidación: {e}\n{traceback.format_exc()}"
            print(error_msg)
            errores.append(error_msg)
    if errores:
        messagebox.showwarning("Advertencia Consolidación", "Ocurrieron errores al consolidar algunos archivos:\n\n" + "\n".join(errores), parent=ventana_principal)
    if not lista_dfs_a_concatenar:
        print("No se consolidaron datos.")
        messagebox.showinfo("Consolidación", "No se pudo consolidar ningún dato de los archivos encontrados.", parent=ventana_principal)
        return None
    df_consolidado = pd.concat(lista_dfs_a_concatenar, ignore_index=True)
    columnas_finales_ordenadas = ['Archivo'] + columnas_interes
    for col in columnas_finales_ordenadas:
        if col not in df_consolidado.columns:
            df_consolidado[col] = pd.NA
    df_consolidado = df_consolidado[columnas_finales_ordenadas]
    try:
        df_consolidado.to_excel(ruta_salida, index=False, engine='openpyxl')
        print(f"Archivo consolidado creado/actualizado: {ruta_salida}")
        return ruta_salida
    except Exception as e:
        error_msg = f"Error al guardar el archivo consolidado '{ruta_salida}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Consolidación", f"No se pudo guardar el archivo consolidado:\n{error_msg}", parent=ventana_principal)
        return None

def _asignar_materiales_interno(ruta_archivo_consolidado):
    import pandas as pd # IMPORTACIÓN DIFERIDA
    # ... (resto del código de _asignar_materiales_interno sin cambios) ...
    print(f"--- Iniciando asignación de materiales en: {ruta_archivo_consolidado} ---")
    if not ruta_archivo_consolidado or not os.path.exists(ruta_archivo_consolidado):
        messagebox.showerror("Error Asignación", f"El archivo consolidado '{os.path.basename(ruta_archivo_consolidado)}' no existe o no se pudo crear.", parent=ventana_principal)
        return False
    mapa_espesores = {16: 'MELAMINA 16 MM', 19: 'EUCALIPTO 18 MM', 20: 'MADERA MELINA', 18: 'MELAMINA 18 MM/EUCALIPTO 18 MM', 
                      22: 'MADERA MELINA', 10: 'MADERA MELINA', 25: 'BUTCHERBLOCK', 3: 'MADERA MELINA', 2: 'MADERA MELINA', 
                      42: 'BUTCHERBLOCK', 50: 'BUTCHERBLOCK', 5: 'MADERA MELINA', 17: 'MELAMINA 18MM', 9: 'PLYWOOD/MADERA DURA'}
    try:
        df = pd.read_excel(ruta_archivo_consolidado, engine='openpyxl')
        if 'Espesor' not in df.columns:
            messagebox.showerror("Error Asignación", "La columna 'Espesor' no se encuentra en el archivo consolidado.", parent=ventana_principal)
            return False
        def convertir_a_int_o_mantener(val):
            if pd.isna(val): return None
            try:
                num = pd.to_numeric(val)
                if pd.notna(num) and num == int(num): return int(num)
                return num
            except (ValueError, TypeError): return val
        df['Espesor_Para_Mapeo'] = df['Espesor'].apply(convertir_a_int_o_mantener)
        df['Material'] = df['Espesor_Para_Mapeo'].map(mapa_espesores).fillna('Material no definido')
        df = df.drop(columns=['Espesor_Para_Mapeo'])
        cols = df.columns.tolist()
        if 'Material' in cols: cols.remove('Material')
        if 'Espesor' in cols:
            idx_espesor = cols.index('Espesor')
            cols.insert(idx_espesor + 1, 'Material')
        else: cols.append('Material')
        df = df[cols]
        df.to_excel(ruta_archivo_consolidado, index=False, engine='openpyxl')
        print(f"Columna 'Material' añadida y archivo guardado: {ruta_archivo_consolidado}")
        return True
    except Exception as e:
        error_msg = f"Error durante la asignación de materiales en '{os.path.basename(ruta_archivo_consolidado)}': {e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Asignación", f"Error al asignar materiales:\n{error_msg}", parent=ventana_principal)
        return False

# --- Funciones de los Botones Principales (sin cambios en lógica interna) ---
# (procesar_y_asignar_flujo_completo, insertar_datos_en_hoja_destino)
# Estas funciones NO cambian respecto a la versión _optimizado.py,
# Solo asegúrate de que las importaciones de pandas/openpyxl están DENTRO de ellas.
def procesar_y_asignar_flujo_completo():
    directorio = seleccionar_carpeta("Selecciona el directorio con los archivos Excel a procesar")
    if not directorio:
        messagebox.showwarning("Cancelado", "No se seleccionó ningún directorio.", parent=ventana_principal)
        return
    nombre_archivo_fusion = "Fusion.xlsx"
    print("\n--- PASO 1: Eliminando 'mm' ---")
    if not _eliminar_mm_interno(directorio, nombre_archivo_fusion):
        print("Proceso detenido debido a errores durante la eliminación de 'mm'.")
        return
    print("\n--- PASO 2: Consolidando datos ---")
    ruta_consolidado = _consolidar_datos_interno(directorio, nombre_archivo_fusion)
    if not ruta_consolidado:
        print(f"Proceso detenido. No se pudo crear o no hay datos para '{nombre_archivo_fusion}'.")
        return
    print("\n--- PASO 3: Asignando materiales ---")
    if _asignar_materiales_interno(ruta_consolidado):
        messagebox.showinfo("Éxito", f"Proceso completado.\n\nSe eliminó 'mm', se consolidaron los datos y se asignaron materiales en:\n{ruta_consolidado}", parent=ventana_principal)
    else:
        print(f"El proceso de asignación de materiales falló. El archivo {ruta_consolidado} puede estar incompleto.")
        messagebox.showerror("Error Final", f"El proceso de asignación de materiales falló.\nRevise el archivo {ruta_consolidado} y los logs de consola.", parent=ventana_principal)

def insertar_datos_en_hoja_destino():
    import pandas as pd # IMPORTACIÓN DIFERIDA
    from openpyxl import load_workbook # IMPORTACIÓN DIFERIDA
    from openpyxl.styles import PatternFill # IMPORTACIÓN DIFERIDA
    # ... (resto del código de insertar_datos_en_hoja_destino sin cambios) ...
    print("\n--- Iniciando inserción en hoja destino ---")
    nombre_archivo_esperado = "Fusion.xlsx"
    source_file = seleccionar_archivo(f"Seleccione el archivo CONSOLIDADO ({nombre_archivo_esperado})", filetypes=[("Archivo Consolidado", nombre_archivo_esperado), ("Archivos Excel", "*.xlsx")])
    if not source_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo fuente.", parent=ventana_principal)
        return
    if os.path.basename(source_file) != nombre_archivo_esperado:
        if not messagebox.askyesno("Confirmar Fuente", f"El archivo seleccionado no se llama '{nombre_archivo_esperado}'.\n({os.path.basename(source_file)})\n\n¿Continuar usando este archivo?", parent=ventana_principal):
            return
    try:
        print(f"Leyendo archivo fuente: {source_file}")
        source_data = pd.read_excel(source_file, engine='openpyxl')
        required_cols = ['QTY', 'DESCRIPTION', 'Largo', 'Ancho', 'Espesor', 'Material', 'Archivo']
        missing_cols = [col for col in required_cols if col not in source_data.columns]
        if missing_cols:
            messagebox.showerror("Error Columnas", f"El archivo fuente '{os.path.basename(source_file)}' no contiene las columnas necesarias: {', '.join(missing_cols)}", parent=ventana_principal)
            return
        print("Columnas requeridas encontradas.")
    except Exception as e:
        messagebox.showerror("Error Lectura Fuente", f"Error al leer el archivo fuente:\n{e}\n{traceback.format_exc()}", parent=ventana_principal)
        return
    destination_file = seleccionar_archivo("Seleccione el archivo Excel DESTINO (con la hoja 'LISTA')")
    if not destination_file:
        messagebox.showwarning("Cancelado", "No se seleccionó el archivo destino.", parent=ventana_principal)
        return
    sheet_name_target = 'LISTA'
    try:
        print(f"Cargando archivo destino: {destination_file}")
        wb = load_workbook(destination_file)
        sheet_name_found = None
        for s_name_iter in wb.sheetnames:
            if s_name_iter.upper() == sheet_name_target.upper():
                sheet_name_found = s_name_iter
                break
        if not sheet_name_found:
             messagebox.showerror("Error Hoja Destino", f"La hoja '{sheet_name_target}' no se encuentra en el archivo destino.", parent=ventana_principal)
             return
        ws = wb[sheet_name_found]
        print(f"Hoja '{sheet_name_found}' encontrada en el archivo destino.")
    except Exception as e:
        messagebox.showerror("Error Carga Destino", f"Error al cargar el archivo destino:\n{e}\n{traceback.format_exc()}", parent=ventana_principal)
        return
    start_row = 13
    print(f"Iniciando inserción de datos en la hoja '{sheet_name_found}' a partir de la fila {start_row}...")
    codigo_map = {'MELAMINA 18 MM': 2, 'MELAMINA 16 MM': 2, 'EUCALIPTO 18 MM': 2, 'MADERA MELINA': 1, 'BUTCHERBLOCK': 1}
    color_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    color_azul = PatternFill(start_color="40A2E3", end_color="40A2E3", fill_type="solid")
    no_color = PatternFill(fill_type=None)
    try:
        for i, row_data in enumerate(source_data.itertuples(index=False, name='Pandas')):
            row_num = start_row + i
            material = getattr(row_data, 'Material', '')
            codigo = codigo_map.get(material, '')
            ws[f'A{row_num}'] = getattr(row_data, 'QTY', '')
            ws[f'B{row_num}'] = getattr(row_data, 'DESCRIPTION', '')
            ws[f'C{row_num}'] = getattr(row_data, 'Largo', '')
            ws[f'D{row_num}'] = getattr(row_data, 'Ancho', '')
            ws[f'E{row_num}'] = getattr(row_data, 'Espesor', '')
            ws[f'F{row_num}'] = codigo
            ws[f'G{row_num}'] = material
            ws[f'H{row_num}'] = getattr(row_data, 'Archivo', '')
            cell_f = ws[f'F{row_num}']
            if codigo == 1: cell_f.fill = color_amarillo
            elif codigo == 2: cell_f.fill = color_azul
            else: cell_f.fill = no_color
            if (i + 1) % 100 == 0: print(f"  Insertadas {i + 1} filas...")
        print(f"Inserción completada. Total de {len(source_data)} filas procesadas.")
        wb.save(destination_file)
        print(f"Archivo destino guardado: {destination_file}")
        messagebox.showinfo("Éxito", f"Datos insertados y formateados correctamente desde\n'{os.path.basename(source_file)}'\nen la hoja '{sheet_name_found}' del archivo\n'{os.path.basename(destination_file)}'.", parent=ventana_principal)
    except Exception as e:
        error_msg = f"Error durante la inserción o guardado:\n{e}\n{traceback.format_exc()}"
        print(error_msg)
        messagebox.showerror("Error Inserción/Guardado", error_msg, parent=ventana_principal)

# --- Configuración de la Interfaz Gráfica Principal con tkinter.ttk ---

ventana_principal = tk.Tk()
ventana_principal.title("Gestor de Listas COPT v2.2 (Ligero)")
ventana_principal.geometry("550x330") # Ajustado ligeramente

# Estilo con ttk.Style
style = ttk.Style(ventana_principal)

# Intentar usar un tema ttk que se vea bien y esté disponible
# 'clam', 'alt', 'default', 'classic'. 'vista' en Windows.
# Puedes experimentar con style.theme_names() para ver los disponibles
available_themes = style.theme_names()
if 'clam' in available_themes:
    style.theme_use('clam')
elif 'vista' in available_themes: # 'vista' suele verse bien en Windows
    style.theme_use('vista')
else: # Si no, el default
    pass # Se usará el tema por defecto del sistema para ttk

# Definir fuentes (ajusta 'Segoe UI' si no está disponible en tu sistema de prueba)
try:
    default_font = tkFont.Font(family="Segoe UI", size=10)
    bold_font = tkFont.Font(family="Segoe UI", size=10, weight="bold")
    link_font = tkFont.Font(family="Segoe UI", size=9, underline=True)
except tk.TclError: # Fallback a fuentes comunes
    default_font = tkFont.Font(family="Arial", size=10)
    bold_font = tkFont.Font(family="Arial", size=10, weight="bold")
    link_font = tkFont.Font(family="Arial", size=9, underline=True)


# Aplicar estilos globales a los widgets ttk
style.configure('.', font=default_font, padding=3) # Padding para todos los widgets ttk
style.configure('TButton', padding=(8, 4), font=default_font) # Padding específico para botones
style.configure('Bold.TLabel', font=bold_font)
style.configure('Link.TLabel', font=link_font, foreground='blue') # Estilo para el label de "link"

# Frame principal para mejor organización y padding general
main_frame = ttk.Frame(ventana_principal, padding=(20, 10))
main_frame.pack(fill=tk.BOTH, expand=True)

# --- Widgets ---
# Botón 1: Procesar
label1 = ttk.Label(main_frame, text="Paso 1: Procesar archivos y generar 'Fusion.xlsx'", style='Bold.TLabel')
label1.pack(pady=(10, 5))
boton_procesar = ttk.Button(main_frame,
                            text="Procesar Carpeta y Asignar Materiales",
                            command=procesar_y_asignar_flujo_completo)
boton_procesar.pack(pady=5, fill=tk.X, padx=10)

# Separador visual
separator1 = ttk.Separator(main_frame, orient=tk.HORIZONTAL)
separator1.pack(pady=15, fill=tk.X, padx=10)

# Botón 2: Insertar
label2 = ttk.Label(main_frame, text="Paso 2: Cargar 'Fusion.xlsx' en la hoja 'LISTA' destino", style='Bold.TLabel')
label2.pack(pady=(5, 5))
boton_insertar = ttk.Button(main_frame,
                            text="Insertar Datos en Hoja Destino",
                            command=insertar_datos_en_hoja_destino)
boton_insertar.pack(pady=5, fill=tk.X, padx=10)

# Separador visual
separator2 = ttk.Separator(main_frame, orient=tk.HORIZONTAL)
separator2.pack(pady=15, fill=tk.X, padx=10)

# Info Desarrollador y Versión
def abrir_ayuda_web(): # Renombrada para evitar conflicto si 'abrir_ayuda' se usara en otro lado
    webbrowser.open("https://www.pentestercr.com/") # Cambia esto si quieres

# Usar un Label para el "link" del desarrollador, con binding
dev_label = ttk.Label(main_frame, text="Desarrollado por S4mma3l", style='Link.TLabel', cursor="hand2")
dev_label.pack(pady=(10, 0))
dev_label.bind("<Button-1>", lambda e: abrir_ayuda_web())

version_label = ttk.Label(main_frame, text="Version: 2.2.0 (Ligero)") # Puedes usar style='Secondary.TLabel' si defines uno
version_label.pack(pady=(0, 10))

# Iniciar el bucle principal de la GUI
ventana_principal.mainloop()

# --- END OF FILE programa_v2.2_ligero.py ---